import json
import pandas as pd
from apify_client import ApifyClient
from openpyxl import Workbook, load_workbook
import logging

logging.basicConfig(
    filename="tiktok_automatizacion.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def obtener_datos_tiktok(cuentas: list, apify_token: str, limite: int = 50) -> dict:
    """
    Utiliza el scraper de Apify para extraer datos de TikTok.
    """
    client = ApifyClient(apify_token)
    datos_por_cuenta = {}

    for cuenta in cuentas:
        try:
            print(f"Extrayendo datos de la cuenta: {cuenta}")
            run = client.actor("0FXVyOXXEmdGcV88a").call(
                run_input={
                    "profiles": [cuenta],
                    "resultsLimit": limite,  
                    "resultsPerPage": limite
                }
            )
            dataset_id = run["defaultDatasetId"]
            dataset_client = client.dataset(dataset_id)
            items = dataset_client.list_items().items
            datos_por_cuenta[cuenta] = items
        except Exception as e:
            logging.error(f"Error extrayendo datos de la cuenta '{cuenta}': {e}")
    
    return datos_por_cuenta

def procesar_datos(datos_por_cuenta: dict) -> dict:
    """
    Procesa los datos extraídos para preparar DataFrames organizados por cuenta.
    """
    dataframes = {}
    for cuenta, videos in datos_por_cuenta.items():
        registros = []
        for video in videos:
            try:
                registros.append({
                    "URL": video.get("webVideoUrl", "Sin URL"),
                    "Visitas": video.get("playCount", 0),
                    "Likes": video.get("diggCount", 0),
                    "Comentarios": video.get("commentCount", 0),
                    "Compartidos": video.get("shareCount", 0),
                    "Fecha": video.get("createTimeISO", "Sin Fecha"),
                })
            except Exception as e:
                logging.error(f"Error procesando video de la cuenta '{cuenta}': {e}")
        
        if registros:
            dataframes[cuenta] = pd.DataFrame(registros)
        else:
            logging.warning(f"No se encontraron datos válidos para la cuenta: {cuenta}")
    return dataframes

def guardar_en_excel(dataframes: dict, nombre_archivo: str):
    """
    Guarda los DataFrames procesados en un archivo Excel.
    """
    try:
        try:
            libro = load_workbook(nombre_archivo)
        except FileNotFoundError:
            libro = Workbook()
            libro.remove(libro.active)  
        
        for cuenta, df in dataframes.items():
            if cuenta in libro.sheetnames:
                hoja = libro[cuenta]
                for row in hoja.iter_rows():
                    for cell in row:
                        cell.value = None
            else:
                hoja = libro.create_sheet(title=cuenta)
            
            encabezados = ["URL", "Visitas", "Likes", "Comentarios", "Compartidos","Fecha"]
            hoja.append(encabezados)

            for _, row in df.iterrows():
                hoja.append(row.tolist())

        libro.save(nombre_archivo)
        logging.info(f"Datos guardados correctamente en {nombre_archivo}")
    except Exception as e:
        logging.error(f"Error al guardar los datos en Excel: {e}")

def main():
    """
    Función principal 
    """
    apify_token = "apify_api_9yE5TuLq4QD9rEI6Z4t1PfTDLhPo663uxcS4"  #TOKEN API SE CAMBIA DEPENDIENDO DE TU CUENTA DE APIFY, ESTE ES EL MIO 
                                                                    #(NO SE PUEDE COMPARTIR PORQUE CON ESE TOKEN TE PUEDE ROBAR LOS DATOS)
    cuentas = ["trapzone36","tt.tendencias","argentinetrap1","trappin.lyrics","555liveforever","brattraplyrics_","555ramma","ara.youngboy","v4luto","gacetaplay"]  
    limite_videos = 50  # Cambia según lo necesario

    # Extraer datos
    datos_por_cuenta = obtener_datos_tiktok(cuentas, apify_token, limite_videos)
    
    # Procesar datos
    dataframes = procesar_datos(datos_por_cuenta)

    # Guardar en Excel
    guardar_en_excel(dataframes, "Datos_TikTok.xlsx")

if __name__ == "__main__":
    main()


    
    
